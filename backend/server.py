from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.encoders import jsonable_encoder
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr, field_serializer
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import pytz
import bcrypt
import jwt
from passlib.context import CryptContext

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Timezone Configuration
IST = pytz.timezone('Asia/Kolkata')

def get_ist_now():
    """Get current time in IST timezone"""
    return datetime.now(IST)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24  # 24 hours

# Custom JSON response to handle timezone-aware datetimes
from fastapi.responses import JSONResponse
from typing import Any
import json

class CustomJSONResponse(JSONResponse):
    def render(self, content: Any) -> bytes:
        return json.dumps(
            content,
            ensure_ascii=False,
            allow_nan=False,
            indent=None,
            separators=(",", ":"),
            default=str  # This will call __str__ on datetime objects, preserving timezone
        ).encode("utf-8")

app = FastAPI(default_response_class=CustomJSONResponse)
api_router = APIRouter(prefix="/api")

# ========== MODELS ==========

class UserRegister(BaseModel):
    username: str
    email: EmailStr
    password: str
    full_name: str

class UserLogin(BaseModel):
    username: str
    password: str

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    full_name: str
    is_admin: bool = False
    created_at: datetime = Field(default_factory=get_ist_now)

class TokenResponse(BaseModel):
    access_token: str
    token_type: str
    user: User

class ProductCreate(BaseModel):
    name: str
    category: str  # chicken, mutton, frozen, liver, kidney
    description: Optional[str] = None
    unit: str  # kg, piece
    price_per_unit: float
    stock_quantity: float
    reorder_level: float
    sku: Optional[str] = None
    is_raw_material: bool = False
    purchase_cost: float = 0.0
    derived_from: Optional[str] = None

class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    category: str
    description: Optional[str] = None
    unit: str
    price_per_unit: float
    stock_quantity: float
    reorder_level: float
    sku: Optional[str] = None
    is_raw_material: bool = False
    purchase_cost: float = 0.0
    derived_from: Optional[str] = None
    created_at: datetime = Field(default_factory=get_ist_now)
    updated_at: datetime = Field(default_factory=get_ist_now)

class VendorCreate(BaseModel):
    name: str
    contact_person: str
    phone: str
    email: Optional[str] = None
    address: Optional[str] = None

class Vendor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    contact_person: str
    phone: str
    email: Optional[str] = None
    address: Optional[str] = None
    created_at: datetime = Field(default_factory=get_ist_now)

class CustomerCreate(BaseModel):
    name: str
    phone: str
    email: Optional[str] = None
    address: Optional[str] = None

class Customer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: str
    email: Optional[str] = None
    address: Optional[str] = None
    total_purchases: float = 0.0
    created_at: datetime = Field(default_factory=get_ist_now)

class SaleItem(BaseModel):
    product_id: str
    product_name: str
    quantity: float
    unit: str
    price_per_unit: float
    total: float

class SaleCreate(BaseModel):
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    items: List[SaleItem]
    subtotal: float
    tax: float
    discount: float
    total: float
    payment_method: str  # cash, card, upi

class Sale(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    items: List[SaleItem]
    subtotal: float
    tax: float
    discount: float
    total: float
    payment_method: str
    created_at: datetime = Field(default_factory=get_ist_now)
    created_by: str

class DashboardStats(BaseModel):
    total_sales_today: float
    total_sales_month: float
    low_stock_items: int
    total_customers: int
    total_products: int
    recent_sales: List[Sale]

class PurchaseCreate(BaseModel):
    vendor_id: str
    raw_material_id: str
    quantity: float
    cost_per_unit: float
    total_cost: float
    purchase_date: Optional[str] = None

class Purchase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    vendor_id: str
    vendor_name: str
    raw_material_id: str
    raw_material_name: str
    quantity: float
    unit: str
    cost_per_unit: float
    total_cost: float
    purchase_date: datetime = Field(default_factory=get_ist_now)
    created_by: str

# ========== AUTHENTICATION ==========

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = get_ist_now() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return User(**user)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception as e:
        raise HTTPException(status_code=401, detail="Invalid token")

# Initialize admin user on startup
@app.on_event("startup")
async def create_admin_user():
    try:
        # Check if admin exists
        admin_exists = await db.users.find_one({"username": "admin-bano"}, {"_id": 0})
        if not admin_exists:
            # Create admin user
            hashed_password = pwd_context.hash("India@54321")
            admin_user = {
                "id": str(uuid.uuid4()),
                "username": "admin-bano",
                "email": "admin@banofresh.com",
                "full_name": "Bano Fresh Admin",
                "is_admin": True,
                "created_at": get_ist_now().isoformat()
            }
            admin_user['password'] = hashed_password
            await db.users.insert_one(admin_user)
            logger.info("✅ Admin user created: admin-bano")
        else:
            logger.info("✅ Admin user already exists")
    except Exception as e:
        logger.error(f"Error creating admin user: {e}")

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(user_input: UserLogin):
    user_doc = await db.users.find_one({"username": user_input.username}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not pwd_context.verify(user_input.password, user_doc['password']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    user = User(**{k: v for k, v in user_doc.items() if k != 'password'})
    access_token = create_access_token({"sub": user.id})
    
    return TokenResponse(
        access_token=access_token,
        token_type="bearer",
        user=user
    )

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

# Admin-only: Create new user
class UserCreate(BaseModel):
    username: str
    email: EmailStr
    password: str
    full_name: str
    is_admin: bool = False


# ========== NEW INVENTORY SYSTEM MODELS ==========

class MainCategoryCreate(BaseModel):
    name: str
    description: Optional[str] = None

class MainCategory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    created_at: datetime = Field(default_factory=get_ist_now)
    updated_at: datetime = Field(default_factory=get_ist_now)

class DerivedProductCreate(BaseModel):
    main_category_id: str
    name: str
    sku: str
    sale_unit: str  # "weight", "package", or "pieces"
    package_weight_kg: Optional[float] = None  # Required if sale_unit = "package"
    selling_price: float
    description: Optional[str] = None

class DerivedProduct(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    main_category_id: str
    name: str
    sku: str
    sale_unit: str = "weight"  # "weight", "package", or "pieces" - default to "weight" for backwards compatibility
    package_weight_kg: Optional[float] = None  # Package weight in kg (e.g., 0.5 for 500g)
    selling_price: float
    description: Optional[str] = None
    created_at: datetime = Field(default_factory=get_ist_now)
    updated_at: datetime = Field(default_factory=get_ist_now)

class InventoryPurchaseCreate(BaseModel):
    main_category_id: str
    vendor_id: str
    total_weight_kg: float
    total_pieces: Optional[int] = None
    cost_per_kg: float
    notes: Optional[str] = None

class InventoryPurchase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    main_category_id: str
    main_category_name: str
    vendor_id: str
    vendor_name: str
    purchase_date: datetime = Field(default_factory=get_ist_now)
    total_weight_kg: float
    total_pieces: Optional[int] = None
    remaining_weight_kg: float
    remaining_pieces: Optional[int] = None
    cost_per_kg: float
    total_cost: float
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=get_ist_now)
    
    @field_serializer('purchase_date', 'created_at')
    def serialize_datetime(self, dt: datetime, _info):
        """Serialize datetime with timezone info"""
        if dt.tzinfo is None:
            dt = IST.localize(dt)
        return dt.isoformat()

class DailyPiecesTrackingCreate(BaseModel):
    main_category_id: str
    pieces_sold: int
    tracking_date: Optional[str] = None  # YYYY-MM-DD format

class DailyPiecesTracking(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    main_category_id: str
    main_category_name: str
    tracking_date: str  # YYYY-MM-DD format
    pieces_sold: int
    created_at: datetime = Field(default_factory=get_ist_now)

class DailyWasteTrackingCreate(BaseModel):
    main_category_id: str
    waste_kg: float  # Direct waste amount in kg
    notes: Optional[str] = None
    tracking_date: Optional[str] = None  # YYYY-MM-DD format

class DailyWasteTracking(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    main_category_id: str
    main_category_name: str
    tracking_date: str  # YYYY-MM-DD format
    waste_kg: float  # Direct waste amount in kg
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=get_ist_now)
    
    @field_serializer('created_at')
    def serialize_datetime(self, dt: datetime, _info):
        """Serialize datetime with timezone info"""
        if dt.tzinfo is None:
            dt = IST.localize(dt)
        return dt.isoformat()

class POSSaleItemNew(BaseModel):
    derived_product_id: str
    derived_product_name: str
    main_category_id: str
    main_category_name: str
    quantity_kg: float  # Weight in kg (for weight/package) or 0 (for pieces)
    quantity_pieces: Optional[int] = None  # Number of pieces (for pieces unit)
    selling_price: float
    total: float

class POSSaleCreateNew(BaseModel):
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    items: List[POSSaleItemNew]
    subtotal: float
    tax: float
    discount: float
    total: float
    payment_method: str

class POSSaleNew(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    items: List[POSSaleItemNew]
    subtotal: float
    tax: float
    discount: float
    total: float
    payment_method: str
    sale_date: datetime = Field(default_factory=get_ist_now)
    created_at: datetime = Field(default_factory=get_ist_now)
    
    @field_serializer('sale_date', 'created_at')
    def serialize_datetime(self, dt: datetime, _info):
        """Serialize datetime with timezone info"""
        if dt.tzinfo is None:
            # If no timezone, assume IST
            dt = IST.localize(dt)
        return dt.isoformat()

class InventorySummary(BaseModel):
    main_category_id: str
    main_category_name: str
    total_weight_kg: float
    total_pieces: int
    low_stock: bool = False
    today_waste_kg: float = 0.0
    today_waste_percentage: float = 0.0
    week_waste_kg: float = 0.0
    week_waste_percentage: float = 0.0

@api_router.post("/users", response_model=User)
async def create_user(user_input: UserCreate, current_user: User = Depends(get_current_user)):
    # Check if current user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can create users")
    
    # Check if user exists
    existing = await db.users.find_one({"$or": [{"username": user_input.username}, {"email": user_input.email}]}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Username or email already exists")
    
    # Hash password
    hashed_password = pwd_context.hash(user_input.password)
    
    # Create user
    user = User(
        username=user_input.username,
        email=user_input.email,
        full_name=user_input.full_name
    )
    
    user_doc = user.model_dump()
    user_doc['password'] = hashed_password
    user_doc['is_admin'] = user_input.is_admin
    user_doc['created_at'] = user_doc['created_at'].isoformat()
    
    await db.users.insert_one(user_doc)
    return user

@api_router.get("/users", response_model=List[User])
async def get_users(current_user: User = Depends(get_current_user)):
    # Check if current user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can view users")
    
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    for u in users:
        if isinstance(u.get('created_at'), str):
            u['created_at'] = datetime.fromisoformat(u['created_at'])
    return users

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    # Check if current user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can delete users")
    
    # Don't allow deleting admin user
    target_user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if target_user and target_user.get('username') == 'admin-bano':
        raise HTTPException(status_code=400, detail="Cannot delete admin user")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted successfully"}

# ========== PRODUCTS ==========

@api_router.post("/products", response_model=Product)
async def create_product(product_input: ProductCreate, current_user: User = Depends(get_current_user)):
    product = Product(**product_input.model_dump())
    doc = product.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.products.insert_one(doc)
    return product

@api_router.get("/products", response_model=List[Product])
async def get_products(current_user: User = Depends(get_current_user)):
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    for p in products:
        if isinstance(p.get('created_at'), str):
            p['created_at'] = datetime.fromisoformat(p['created_at'])
        if isinstance(p.get('updated_at'), str):
            p['updated_at'] = datetime.fromisoformat(p['updated_at'])
    return products

@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str, current_user: User = Depends(get_current_user)):
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    if isinstance(product.get('created_at'), str):
        product['created_at'] = datetime.fromisoformat(product['created_at'])
    if isinstance(product.get('updated_at'), str):
        product['updated_at'] = datetime.fromisoformat(product['updated_at'])
    return Product(**product)

@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(product_id: str, product_input: ProductCreate, current_user: User = Depends(get_current_user)):
    existing = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")
    
    update_data = product_input.model_dump()
    update_data['updated_at'] = get_ist_now().isoformat()
    
    await db.products.update_one({"id": product_id}, {"$set": update_data})
    
    updated = await db.products.find_one({"id": product_id}, {"_id": 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    if isinstance(updated.get('updated_at'), str):
        updated['updated_at'] = datetime.fromisoformat(updated['updated_at'])
    return Product(**updated)

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: User = Depends(get_current_user)):
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product deleted successfully"}

@api_router.get("/products/low-stock/alert", response_model=List[Product])
async def get_low_stock_products(current_user: User = Depends(get_current_user)):
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    low_stock = []
    for p in products:
        if p['stock_quantity'] <= p['reorder_level']:
            if isinstance(p.get('created_at'), str):
                p['created_at'] = datetime.fromisoformat(p['created_at'])
            if isinstance(p.get('updated_at'), str):
                p['updated_at'] = datetime.fromisoformat(p['updated_at'])
            low_stock.append(Product(**p))
    return low_stock

# ========== VENDORS ==========

@api_router.post("/vendors", response_model=Vendor)
async def create_vendor(vendor_input: VendorCreate, current_user: User = Depends(get_current_user)):
    vendor = Vendor(**vendor_input.model_dump())
    doc = vendor.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.vendors.insert_one(doc)
    return vendor

@api_router.get("/vendors", response_model=List[Vendor])
async def get_vendors(current_user: User = Depends(get_current_user)):
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(1000)
    for v in vendors:
        if isinstance(v.get('created_at'), str):
            v['created_at'] = datetime.fromisoformat(v['created_at'])
    return vendors

@api_router.put("/vendors/{vendor_id}", response_model=Vendor)
async def update_vendor(vendor_id: str, vendor_input: VendorCreate, current_user: User = Depends(get_current_user)):
    existing = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    await db.vendors.update_one({"id": vendor_id}, {"$set": vendor_input.model_dump()})
    updated = await db.vendors.find_one({"id": vendor_id}, {"_id": 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Vendor(**updated)

@api_router.delete("/vendors/{vendor_id}")
async def delete_vendor(vendor_id: str, current_user: User = Depends(get_current_user)):
    result = await db.vendors.delete_one({"id": vendor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return {"message": "Vendor deleted successfully"}

# ========== CUSTOMERS ==========

@api_router.post("/customers", response_model=Customer)
async def create_customer(customer_input: CustomerCreate, current_user: User = Depends(get_current_user)):
    customer = Customer(**customer_input.model_dump())
    doc = customer.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.customers.insert_one(doc)
    return customer

@api_router.get("/customers", response_model=List[Customer])
async def get_customers(current_user: User = Depends(get_current_user)):
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    for c in customers:
        if isinstance(c.get('created_at'), str):
            c['created_at'] = datetime.fromisoformat(c['created_at'])
    return customers

@api_router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, customer_input: CustomerCreate, current_user: User = Depends(get_current_user)):
    existing = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    await db.customers.update_one({"id": customer_id}, {"$set": customer_input.model_dump()})
    updated = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Customer(**updated)

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: User = Depends(get_current_user)):
    result = await db.customers.delete_one({"id": customer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Customer deleted successfully"}

# ========== SALES ==========

@api_router.post("/sales", response_model=Sale)
async def create_sale(sale_input: SaleCreate, current_user: User = Depends(get_current_user)):
    # Update product stock
    for item in sale_input.items:
        product = await db.products.find_one({"id": item.product_id}, {"_id": 0})
        if product:
            new_stock = product['stock_quantity'] - item.quantity
            await db.products.update_one(
                {"id": item.product_id},
                {"$set": {"stock_quantity": new_stock}}
            )
    
    # Update customer total purchases if customer exists
    if sale_input.customer_id:
        customer = await db.customers.find_one({"id": sale_input.customer_id}, {"_id": 0})
        if customer:
            new_total = customer.get('total_purchases', 0) + sale_input.total
            await db.customers.update_one(
                {"id": sale_input.customer_id},
                {"$set": {"total_purchases": new_total}}
            )
    
    sale = Sale(
        **sale_input.model_dump(),
        created_by=current_user.id
    )
    doc = sale.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['items'] = [item.model_dump() for item in sale.items]
    
    await db.sales.insert_one(doc)
    return sale

@api_router.get("/sales", response_model=List[Sale])
async def get_sales(current_user: User = Depends(get_current_user)):
    sales = await db.sales.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for s in sales:
        if isinstance(s.get('created_at'), str):
            s['created_at'] = datetime.fromisoformat(s['created_at'])
        s['items'] = [SaleItem(**item) for item in s['items']]
    return sales

@api_router.get("/sales/{sale_id}", response_model=Sale)
async def get_sale(sale_id: str, current_user: User = Depends(get_current_user)):
    sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    if isinstance(sale.get('created_at'), str):
        sale['created_at'] = datetime.fromisoformat(sale['created_at'])
    sale['items'] = [SaleItem(**item) for item in sale['items']]
    return Sale(**sale)

# ========== DASHBOARD ==========

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(current_user: User = Depends(get_current_user)):
    # Today's sales
    today_start = get_ist_now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_sales = await db.sales.find({}, {"_id": 0}).to_list(10000)
    total_today = sum(s['total'] for s in today_sales if datetime.fromisoformat(s['created_at']) >= today_start)
    
    # Month's sales
    month_start = get_ist_now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    total_month = sum(s['total'] for s in today_sales if datetime.fromisoformat(s['created_at']) >= month_start)
    
    # Low stock items
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    low_stock_count = sum(1 for p in products if p['stock_quantity'] <= p['reorder_level'])
    
    # Counts
    total_customers = await db.customers.count_documents({})
    total_products = await db.products.count_documents({})
    
    # Recent sales
    recent = await db.sales.find({}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    for s in recent:
        if isinstance(s.get('created_at'), str):
            s['created_at'] = datetime.fromisoformat(s['created_at'])
        s['items'] = [SaleItem(**item) for item in s['items']]
    
    return DashboardStats(
        total_sales_today=total_today,
        total_sales_month=total_month,
        low_stock_items=low_stock_count,
        total_customers=total_customers,
        total_products=total_products,
        recent_sales=[Sale(**s) for s in recent]
    )

# ========== PURCHASES ==========

@api_router.post("/purchases", response_model=Purchase)
async def create_purchase(purchase_input: PurchaseCreate, current_user: User = Depends(get_current_user)):
    # Get vendor and raw material details
    vendor = await db.vendors.find_one({"id": purchase_input.vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    raw_material = await db.products.find_one({"id": purchase_input.raw_material_id}, {"_id": 0})
    if not raw_material:
        raise HTTPException(status_code=404, detail="Raw material not found")
    
    if not raw_material.get('is_raw_material', False):
        raise HTTPException(status_code=400, detail="Selected product is not a raw material")
    
    # Update raw material stock
    new_stock = raw_material['stock_quantity'] + purchase_input.quantity
    await db.products.update_one(
        {"id": purchase_input.raw_material_id},
        {"$set": {"stock_quantity": new_stock, "purchase_cost": purchase_input.cost_per_unit}}
    )
    
    # Create purchase record
    purchase = Purchase(
        vendor_id=purchase_input.vendor_id,
        vendor_name=vendor['name'],
        raw_material_id=purchase_input.raw_material_id,
        raw_material_name=raw_material['name'],
        quantity=purchase_input.quantity,
        unit=raw_material['unit'],
        cost_per_unit=purchase_input.cost_per_unit,
        total_cost=purchase_input.total_cost,
        created_by=current_user.id
    )
    
    if purchase_input.purchase_date:
        purchase.purchase_date = datetime.fromisoformat(purchase_input.purchase_date)
    
    doc = purchase.model_dump()
    doc['purchase_date'] = doc['purchase_date'].isoformat()
    
    await db.purchases.insert_one(doc)
    return purchase

@api_router.get("/purchases", response_model=List[Purchase])
async def get_purchases(current_user: User = Depends(get_current_user)):
    purchases = await db.purchases.find({}, {"_id": 0}).sort("purchase_date", -1).to_list(1000)
    for p in purchases:
        if isinstance(p.get('purchase_date'), str):
            p['purchase_date'] = datetime.fromisoformat(p['purchase_date'])
    return purchases

@api_router.delete("/purchases/{purchase_id}")
async def delete_purchase(purchase_id: str, current_user: User = Depends(get_current_user)):
    result = await db.purchases.delete_one({"id": purchase_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Purchase not found")
    return {"message": "Purchase deleted successfully"}

# ========== REPORTS ==========

from fastapi.responses import StreamingResponse
from io import BytesIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch

@api_router.get("/reports/sales")
async def get_sales_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "json",
    current_user: User = Depends(get_current_user)
):
    # Fetch sales
    sales = await db.sales.find({}, {"_id": 0}).to_list(10000)
    
    # Filter by date if provided
    if start_date:
        start = datetime.fromisoformat(start_date)
        sales = [s for s in sales if datetime.fromisoformat(s['created_at']) >= start]
    if end_date:
        end = datetime.fromisoformat(end_date).replace(hour=23, minute=59, second=59)
        sales = [s for s in sales if datetime.fromisoformat(s['created_at']) <= end]
    
    if format == "csv":
        output = BytesIO()
        writer = csv.writer(output)
        writer.writerow(['Date', 'Customer', 'Items', 'Subtotal', 'Tax', 'Discount', 'Total', 'Payment Method'])
        for sale in sales:
            writer.writerow([
                sale['created_at'],
                sale.get('customer_name', 'Walk-in'),
                len(sale['items']),
                sale['subtotal'],
                sale['tax'],
                sale['discount'],
                sale['total'],
                sale['payment_method']
            ])
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=sales_report.csv"}
        )
    
    elif format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        
        # Headers
        headers = ['Date', 'Customer', 'Items', 'Subtotal', 'Tax', 'Discount', 'Total', 'Payment Method']
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for sale in sales:
            ws.append([
                sale['created_at'],
                sale.get('customer_name', 'Walk-in'),
                len(sale['items']),
                sale['subtotal'],
                sale['tax'],
                sale['discount'],
                sale['total'],
                sale['payment_method']
            ])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=sales_report.xlsx"}
        )
    
    elif format == "pdf":
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        title = Paragraph("<b>Sales Report</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        data = [['Date', 'Customer', 'Items', 'Subtotal', 'Tax', 'Discount', 'Total', 'Payment']]
        for sale in sales:
            data.append([
                sale['created_at'][:10],
                sale.get('customer_name', 'Walk-in')[:15],
                str(len(sale['items'])),
                f"₹{sale['subtotal']:.2f}",
                f"₹{sale['tax']:.2f}",
                f"₹{sale['discount']:.2f}",
                f"₹{sale['total']:.2f}",
                sale['payment_method']
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=sales_report.pdf"}
        )
    
    else:  # json
        return {"sales": sales, "total_count": len(sales), "total_revenue": sum(s['total'] for s in sales)}

@api_router.get("/reports/inventory")
async def get_inventory_report(format: str = "json", current_user: User = Depends(get_current_user)):
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    
    if format == "csv":
        output = BytesIO()
        writer = csv.writer(output)
        writer.writerow(['Name', 'Category', 'Type', 'Stock', 'Unit', 'Reorder Level', 'Price', 'Purchase Cost', 'Status'])
        for product in products:
            status = 'Low Stock' if product['stock_quantity'] <= product['reorder_level'] else 'In Stock'
            ptype = 'Raw Material' if product.get('is_raw_material', False) else 'Derived Product'
            writer.writerow([
                product['name'],
                product['category'],
                ptype,
                product['stock_quantity'],
                product['unit'],
                product['reorder_level'],
                product['price_per_unit'],
                product.get('purchase_cost', 0),
                status
            ])
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=inventory_report.csv"}
        )
    
    elif format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Report"
        
        headers = ['Name', 'Category', 'Type', 'Stock', 'Unit', 'Reorder Level', 'Price', 'Purchase Cost', 'Status']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for product in products:
            status = 'Low Stock' if product['stock_quantity'] <= product['reorder_level'] else 'In Stock'
            ptype = 'Raw Material' if product.get('is_raw_material', False) else 'Derived Product'
            ws.append([
                product['name'],
                product['category'],
                ptype,
                product['stock_quantity'],
                product['unit'],
                product['reorder_level'],
                product['price_per_unit'],
                product.get('purchase_cost', 0),
                status
            ])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=inventory_report.xlsx"}
        )
    
    elif format == "pdf":
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        styles = getSampleStyleSheet()
        title = Paragraph("<b>Inventory Report</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        data = [['Name', 'Category', 'Type', 'Stock', 'Unit', 'Price', 'Status']]
        for product in products:
            status = 'Low' if product['stock_quantity'] <= product['reorder_level'] else 'OK'
            ptype = 'Raw' if product.get('is_raw_material', False) else 'Derived'
            data.append([
                product['name'][:20],
                product['category'][:10],
                ptype,
                str(product['stock_quantity']),
                product['unit'],
                f"₹{product['price_per_unit']:.0f}",
                status
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#008000')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=inventory_report.pdf"}
        )
    
    else:
        low_stock = [p for p in products if p['stock_quantity'] <= p['reorder_level']]
        return {
            "products": products,
            "total_products": len(products),
            "low_stock_count": len(low_stock),
            "low_stock_items": low_stock
        }

@api_router.get("/reports/purchases")
async def get_purchase_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "json",
    current_user: User = Depends(get_current_user)
):
    purchases = await db.purchases.find({}, {"_id": 0}).to_list(10000)
    
    if start_date:
        start = datetime.fromisoformat(start_date)
        purchases = [p for p in purchases if datetime.fromisoformat(p['purchase_date']) >= start]
    if end_date:
        end = datetime.fromisoformat(end_date).replace(hour=23, minute=59, second=59)
        purchases = [p for p in purchases if datetime.fromisoformat(p['purchase_date']) <= end]
    
    if format == "csv":
        output = BytesIO()
        writer = csv.writer(output)
        writer.writerow(['Date', 'Vendor', 'Raw Material', 'Quantity', 'Unit', 'Cost/Unit', 'Total Cost'])
        for purchase in purchases:
            writer.writerow([
                purchase['purchase_date'],
                purchase['vendor_name'],
                purchase['raw_material_name'],
                purchase['quantity'],
                purchase['unit'],
                purchase['cost_per_unit'],
                purchase['total_cost']
            ])
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=purchase_report.csv"}
        )
    
    elif format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchase Report"
        
        headers = ['Date', 'Vendor', 'Raw Material', 'Quantity', 'Unit', 'Cost/Unit', 'Total Cost']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for purchase in purchases:
            ws.append([
                purchase['purchase_date'],
                purchase['vendor_name'],
                purchase['raw_material_name'],
                purchase['quantity'],
                purchase['unit'],
                purchase['cost_per_unit'],
                purchase['total_cost']
            ])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=purchase_report.xlsx"}
        )
    
    elif format == "pdf":
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        title = Paragraph("<b>Purchase Report</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        data = [['Date', 'Vendor', 'Raw Material', 'Quantity', 'Unit', 'Cost/Unit', 'Total']]
        for purchase in purchases:
            data.append([
                purchase['purchase_date'][:10],
                purchase['vendor_name'][:20],
                purchase['raw_material_name'][:15],
                str(purchase['quantity']),
                purchase['unit'],
                f"₹{purchase['cost_per_unit']:.2f}",
                f"₹{purchase['total_cost']:.2f}"
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF6600')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=purchase_report.pdf"}
        )
    
    else:
        return {
            "purchases": purchases,
            "total_count": len(purchases),
            "total_cost": sum(p['total_cost'] for p in purchases)
        }

@api_router.get("/reports/profit-loss")
async def get_profit_loss_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    format: str = "json",
    current_user: User = Depends(get_current_user)
):
    # Get sales and purchases
    sales = await db.sales.find({}, {"_id": 0}).to_list(10000)
    purchases = await db.purchases.find({}, {"_id": 0}).to_list(10000)
    
    # Filter by date
    if start_date:
        start = datetime.fromisoformat(start_date)
        sales = [s for s in sales if datetime.fromisoformat(s['created_at']) >= start]
        purchases = [p for p in purchases if datetime.fromisoformat(p['purchase_date']) >= start]
    if end_date:
        end = datetime.fromisoformat(end_date).replace(hour=23, minute=59, second=59)
        sales = [s for s in sales if datetime.fromisoformat(s['created_at']) <= end]
        purchases = [p for p in purchases if datetime.fromisoformat(p['purchase_date']) <= end]
    
    total_revenue = sum(s['total'] for s in sales)
    total_purchase_cost = sum(p['total_cost'] for p in purchases)
    gross_profit = total_revenue - total_purchase_cost
    profit_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    if format == "csv":
        output = BytesIO()
        writer = csv.writer(output)
        writer.writerow(['Metric', 'Amount'])
        writer.writerow(['Total Revenue', total_revenue])
        writer.writerow(['Total Purchase Cost', total_purchase_cost])
        writer.writerow(['Gross Profit', gross_profit])
        writer.writerow(['Profit Margin %', f"{profit_margin:.2f}%"])
        writer.writerow([])
        writer.writerow(['Sales Count', len(sales)])
        writer.writerow(['Purchase Count', len(purchases)])
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=profit_loss_report.csv"}
        )
    
    elif format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Profit & Loss"
        
        ws['A1'] = 'Profit & Loss Report'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')
        
        ws.append([])
        ws.append(['Metric', 'Amount'])
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)
        
        ws.append(['Total Revenue', total_revenue])
        ws.append(['Total Purchase Cost', total_purchase_cost])
        ws.append(['Gross Profit', gross_profit])
        ws.append(['Profit Margin %', f"{profit_margin:.2f}%"])
        ws.append([])
        ws.append(['Sales Count', len(sales)])
        ws.append(['Purchase Count', len(purchases)])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=profit_loss_report.xlsx"}
        )
    
    elif format == "pdf":
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        styles = getSampleStyleSheet()
        title = Paragraph("<b>Profit & Loss Report</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.5*inch))
        
        data = [
            ['Metric', 'Amount'],
            ['Total Revenue', f"₹{total_revenue:.2f}"],
            ['Total Purchase Cost', f"₹{total_purchase_cost:.2f}"],
            ['Gross Profit', f"₹{gross_profit:.2f}"],
            ['Profit Margin', f"{profit_margin:.2f}%"],
            ['', ''],
            ['Sales Count', str(len(sales))],
            ['Purchase Count', str(len(purchases))]
        ]
        
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6600CC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#CCFFCC') if gross_profit > 0 else colors.HexColor('#FFCCCC'))
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=profit_loss_report.pdf"}
        )
    
    else:
        return {
            "total_revenue": total_revenue,
            "total_purchase_cost": total_purchase_cost,
            "gross_profit": gross_profit,
            "profit_margin": profit_margin,
            "sales_count": len(sales),
            "purchase_count": len(purchases)
        }

# ========== ROOT ==========


# ========== NEW INVENTORY SYSTEM ENDPOINTS ==========

# Main Categories Management
@api_router.get("/main-categories", response_model=List[MainCategory])
async def get_main_categories(current_user: User = Depends(get_current_user)):
    categories = await db.main_categories.find({}, {"_id": 0}).to_list(length=None)
    return categories

@api_router.post("/main-categories", response_model=MainCategory)
async def create_main_category(category: MainCategoryCreate, current_user: User = Depends(get_current_user)):
    # Check if user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can create main categories")
    
    # Check if category already exists
    existing = await db.main_categories.find_one({"name": category.name}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Category with this name already exists")
    
    new_category = MainCategory(**category.dict())
    await db.main_categories.insert_one(new_category.dict())
    logger.info(f"Main category created: {new_category.name}")
    return new_category

@api_router.put("/main-categories/{category_id}", response_model=MainCategory)
async def update_main_category(category_id: str, category: MainCategoryCreate, current_user: User = Depends(get_current_user)):
    # Check if user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can update main categories")
    
    existing = await db.main_categories.find_one({"id": category_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Category not found")
    
    update_data = category.dict()
    update_data["updated_at"] = get_ist_now()
    
    await db.main_categories.update_one({"id": category_id}, {"$set": update_data})
    
    updated = await db.main_categories.find_one({"id": category_id}, {"_id": 0})
    return MainCategory(**updated)

@api_router.delete("/main-categories/{category_id}")
async def delete_main_category(category_id: str, current_user: User = Depends(get_current_user)):
    # Check if user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can delete main categories")
    
    # Check if category has derived products
    derived_count = await db.derived_products.count_documents({"main_category_id": category_id})
    if derived_count > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete category with {derived_count} derived products")
    
    result = await db.main_categories.delete_one({"id": category_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    
    return {"message": "Category deleted successfully"}

# Derived Products Management
@api_router.get("/derived-products", response_model=List[DerivedProduct])
async def get_derived_products(main_category_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if main_category_id:
        query["main_category_id"] = main_category_id
    
    products = await db.derived_products.find(query, {"_id": 0}).to_list(length=None)
    # Ensure backwards compatibility - add default sale_unit if missing
    for product in products:
        if 'sale_unit' not in product or not product['sale_unit']:
            product['sale_unit'] = 'weight'
        if 'package_weight_kg' not in product:
            product['package_weight_kg'] = None
    return products

@api_router.post("/derived-products", response_model=DerivedProduct)
async def create_derived_product(product: DerivedProductCreate, current_user: User = Depends(get_current_user)):
    # Check if user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can create derived products")
    
    # Check if main category exists
    category = await db.main_categories.find_one({"id": product.main_category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Main category not found")
    
    # Validate sale_unit
    if product.sale_unit not in ["weight", "package", "pieces"]:
        raise HTTPException(status_code=400, detail="sale_unit must be 'weight', 'package', or 'pieces'")
    
    # Validate pieces unit only for Mutton and Frozen
    if product.sale_unit == "pieces":
        category_name = category["name"].lower()
        if "mutton" not in category_name and "frozen" not in category_name:
            raise HTTPException(status_code=400, detail="Pieces unit is only allowed for Mutton and Frozen categories")
    
    # Validate package_weight_kg for package unit
    if product.sale_unit == "package":
        if not product.package_weight_kg or product.package_weight_kg <= 0:
            raise HTTPException(status_code=400, detail="package_weight_kg is required and must be > 0 for package unit")
    
    # Check if SKU already exists
    existing_sku = await db.derived_products.find_one({"sku": product.sku}, {"_id": 0})
    if existing_sku:
        raise HTTPException(status_code=400, detail="Product with this SKU already exists")
    
    new_product = DerivedProduct(**product.dict())
    await db.derived_products.insert_one(new_product.dict())
    logger.info(f"Derived product created: {new_product.name} (SKU: {new_product.sku}, Unit: {new_product.sale_unit})")
    return new_product

@api_router.put("/derived-products/{product_id}", response_model=DerivedProduct)
async def update_derived_product(product_id: str, product: DerivedProductCreate, current_user: User = Depends(get_current_user)):
    # Check if user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can update derived products")
    
    existing = await db.derived_products.find_one({"id": product_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Check if SKU is being changed and if new SKU already exists
    if product.sku != existing.get("sku"):
        existing_sku = await db.derived_products.find_one({"sku": product.sku, "id": {"$ne": product_id}}, {"_id": 0})
        if existing_sku:
            raise HTTPException(status_code=400, detail="Product with this SKU already exists")
    
    update_data = product.dict()
    update_data["updated_at"] = get_ist_now()
    
    await db.derived_products.update_one({"id": product_id}, {"$set": update_data})
    
    updated = await db.derived_products.find_one({"id": product_id}, {"_id": 0})
    return DerivedProduct(**updated)

@api_router.delete("/derived-products/{product_id}")
async def delete_derived_product(product_id: str, current_user: User = Depends(get_current_user)):
    # Check if user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can delete derived products")
    
    result = await db.derived_products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return {"message": "Product deleted successfully"}

# Inventory Purchases Management
@api_router.get("/inventory-purchases", response_model=List[InventoryPurchase])
async def get_inventory_purchases(main_category_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if main_category_id:
        query["main_category_id"] = main_category_id
    
    purchases = await db.inventory_purchases.find(query, {"_id": 0}).sort("purchase_date", -1).to_list(length=None)
    return purchases

@api_router.post("/inventory-purchases", response_model=InventoryPurchase)
async def create_inventory_purchase(purchase: InventoryPurchaseCreate, current_user: User = Depends(get_current_user)):
    # Get main category
    category = await db.main_categories.find_one({"id": purchase.main_category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Main category not found")
    
    # Get vendor
    vendor = await db.vendors.find_one({"id": purchase.vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Calculate total cost
    total_cost = purchase.total_weight_kg * purchase.cost_per_kg
    
    new_purchase = InventoryPurchase(
        **purchase.dict(),
        main_category_name=category["name"],
        vendor_name=vendor["name"],
        remaining_weight_kg=purchase.total_weight_kg,
        remaining_pieces=purchase.total_pieces,
        total_cost=total_cost
    )
    
    await db.inventory_purchases.insert_one(new_purchase.dict())
    logger.info(f"Inventory purchase created: {category['name']} - {purchase.total_weight_kg}kg from {vendor['name']}")
    return new_purchase

@api_router.get("/inventory-summary", response_model=List[InventorySummary])
async def get_inventory_summary(current_user: User = Depends(get_current_user)):
    # Get all main categories
    categories = await db.main_categories.find({}, {"_id": 0}).to_list(length=None)
    
    # Get current date in IST
    today = get_ist_now().strftime("%Y-%m-%d")
    week_ago = (get_ist_now() - timedelta(days=7)).strftime("%Y-%m-%d")
    
    summary = []
    for category in categories:
        # Get all purchases for this category
        purchases = await db.inventory_purchases.find(
            {"main_category_id": category["id"]},
            {"_id": 0}
        ).to_list(length=None)
        
        total_weight = sum(p.get("remaining_weight_kg", 0) for p in purchases)
        total_pieces = sum(p.get("remaining_pieces", 0) for p in purchases if p.get("remaining_pieces") is not None)
        
        # Get today's waste (simplified - just waste_kg)
        today_waste = await db.daily_waste_tracking.find(
            {"main_category_id": category["id"], "tracking_date": today},
            {"_id": 0}
        ).to_list(length=None)
        
        today_waste_kg = sum(w.get("waste_kg", 0) for w in today_waste)
        
        # Get this week's waste
        week_waste = await db.daily_waste_tracking.find(
            {
                "main_category_id": category["id"],
                "tracking_date": {"$gte": week_ago, "$lte": today}
            },
            {"_id": 0}
        ).to_list(length=None)
        
        week_waste_kg = sum(w.get("waste_kg", 0) for w in week_waste)
        
        summary.append(InventorySummary(
            main_category_id=category["id"],
            main_category_name=category["name"],
            total_weight_kg=round(total_weight, 2),
            total_pieces=total_pieces,
            low_stock=total_weight < 10,  # Alert if less than 10kg
            today_waste_kg=round(today_waste_kg, 2),
            today_waste_percentage=0,  # Removed percentage calculation
            week_waste_kg=round(week_waste_kg, 2),
            week_waste_percentage=0  # Removed percentage calculation
        ))
    
    return summary

@api_router.put("/inventory-purchases/{purchase_id}")
async def update_inventory_purchase(
    purchase_id: str,
    update_data: InventoryPurchaseCreate,
    current_user: User = Depends(get_current_user)
):
    # Check if user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can edit purchases")
    
    # Get existing purchase
    existing_purchase = await db.inventory_purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not existing_purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    # Get category and vendor
    category = await db.main_categories.find_one({"id": update_data.main_category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Main category not found")
    
    vendor = await db.vendors.find_one({"id": update_data.vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Calculate the difference in weights/pieces to adjust inventory
    old_total_weight = existing_purchase.get("total_weight_kg", 0)
    old_remaining_weight = existing_purchase.get("remaining_weight_kg", 0)
    old_total_pieces = existing_purchase.get("total_pieces", 0) or 0
    old_remaining_pieces = existing_purchase.get("remaining_pieces", 0) or 0
    
    new_total_weight = update_data.total_weight_kg
    new_total_pieces = update_data.total_pieces or 0
    
    # Calculate how much has been used
    used_weight = old_total_weight - old_remaining_weight
    used_pieces = old_total_pieces - old_remaining_pieces
    
    # New remaining = new_total - used
    new_remaining_weight = max(0, new_total_weight - used_weight)
    new_remaining_pieces = max(0, new_total_pieces - used_pieces) if new_total_pieces > 0 else None
    
    # Update purchase
    total_cost = update_data.total_weight_kg * update_data.cost_per_kg
    
    await db.inventory_purchases.update_one(
        {"id": purchase_id},
        {"$set": {
            "main_category_id": update_data.main_category_id,
            "main_category_name": category["name"],
            "vendor_id": update_data.vendor_id,
            "vendor_name": vendor["name"],
            "total_weight_kg": new_total_weight,
            "total_pieces": new_total_pieces,
            "remaining_weight_kg": round(new_remaining_weight, 2),
            "remaining_pieces": new_remaining_pieces,
            "cost_per_kg": update_data.cost_per_kg,
            "total_cost": round(total_cost, 2),
            "notes": update_data.notes
        }}
    )
    
    updated_purchase = await db.inventory_purchases.find_one({"id": purchase_id}, {"_id": 0})
    logger.info(f"Purchase updated: {purchase_id}")
    return InventoryPurchase(**updated_purchase)

@api_router.delete("/inventory-purchases/{purchase_id}")
async def delete_inventory_purchase(purchase_id: str, current_user: User = Depends(get_current_user)):
    # Check if user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can delete purchases")
    
    # Get existing purchase
    existing_purchase = await db.inventory_purchases.find_one({"id": purchase_id}, {"_id": 0})
    if not existing_purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    
    # Check if purchase has been partially used
    remaining_weight = existing_purchase.get("remaining_weight_kg", 0)
    total_weight = existing_purchase.get("total_weight_kg", 0)
    
    if remaining_weight < total_weight:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete purchase that has been partially used. Remaining: {remaining_weight}kg, Total: {total_weight}kg"
        )
    
    # Delete purchase
    await db.inventory_purchases.delete_one({"id": purchase_id})
    logger.info(f"Purchase deleted: {purchase_id}")
    return {"message": "Purchase deleted successfully"}

# Stock Alerts
@api_router.get("/stock-alerts")
async def get_stock_alerts(current_user: User = Depends(get_current_user)):
    # Get all main categories
    categories = await db.main_categories.find({}, {"_id": 0}).to_list(length=None)
    
    alerts = []
    for category in categories:
        # Get all purchases for this category
        purchases = await db.inventory_purchases.find(
            {"main_category_id": category["id"]},
            {"_id": 0}
        ).to_list(length=None)
        
        total_weight = sum(p.get("remaining_weight_kg", 0) for p in purchases)
        
        # Low stock if less than 10kg
        if total_weight < 10:
            alerts.append({
                "category_id": category["id"],
                "category_name": category["name"],
                "current_stock_kg": round(total_weight, 2),
                "alert_level": "critical" if total_weight < 5 else "warning",
                "message": f"Low stock alert: Only {round(total_weight, 2)}kg remaining"
            })
    
    return alerts

# Daily Pieces Tracking
@api_router.get("/daily-pieces-tracking", response_model=List[DailyPiecesTracking])
async def get_daily_pieces_tracking(
    main_category_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if main_category_id:
        query["main_category_id"] = main_category_id
    if start_date and end_date:
        query["tracking_date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["tracking_date"] = {"$gte": start_date}
    
    tracking = await db.daily_pieces_tracking.find(query, {"_id": 0}).sort("tracking_date", -1).to_list(length=None)
    return tracking

@api_router.post("/daily-pieces-tracking", response_model=DailyPiecesTracking)
async def create_daily_pieces_tracking(tracking: DailyPiecesTrackingCreate, current_user: User = Depends(get_current_user)):
    # Get main category
    category = await db.main_categories.find_one({"id": tracking.main_category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Main category not found")
    
    # Determine tracking date (use IST)
    tracking_date = tracking.tracking_date if tracking.tracking_date else get_ist_now().strftime("%Y-%m-%d")
    
    # Check if already tracked for this date and category
    existing = await db.daily_pieces_tracking.find_one({
        "main_category_id": tracking.main_category_id,
        "tracking_date": tracking_date
    }, {"_id": 0})
    
    if existing:
        raise HTTPException(status_code=400, detail="Pieces already tracked for this category and date. Please update instead.")
    
    # Get all inventory purchases for this category and deduct pieces
    purchases = await db.inventory_purchases.find(
        {"main_category_id": tracking.main_category_id},
        {"_id": 0}
    ).sort("purchase_date", 1).to_list(length=None)
    
    pieces_to_deduct = tracking.pieces_sold
    for purchase in purchases:
        if pieces_to_deduct <= 0:
            break
        
        remaining = purchase.get("remaining_pieces", 0)
        if remaining and remaining > 0:
            deduction = min(remaining, pieces_to_deduct)
            new_remaining = remaining - deduction
            pieces_to_deduct -= deduction
            
            await db.inventory_purchases.update_one(
                {"id": purchase["id"]},
                {"$set": {"remaining_pieces": new_remaining}}
            )
    
    if pieces_to_deduct > 0:
        logger.warning(f"Not enough pieces in inventory. {pieces_to_deduct} pieces could not be deducted.")
    
    new_tracking = DailyPiecesTracking(
        main_category_id=tracking.main_category_id,
        main_category_name=category["name"],
        tracking_date=tracking_date,
        pieces_sold=tracking.pieces_sold
    )
    
    await db.daily_pieces_tracking.insert_one(new_tracking.dict())
    logger.info(f"Daily pieces tracking created: {category['name']} - {tracking.pieces_sold} pieces on {tracking_date}")
    return new_tracking

@api_router.put("/daily-pieces-tracking/{tracking_id}")
async def update_daily_pieces_tracking(
    tracking_id: str,
    update_data: DailyPiecesTrackingCreate,
    current_user: User = Depends(get_current_user)
):
    # Check if user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can edit daily pieces tracking")
    
    # Get existing tracking
    existing_tracking = await db.daily_pieces_tracking.find_one({"id": tracking_id}, {"_id": 0})
    if not existing_tracking:
        raise HTTPException(status_code=404, detail="Tracking record not found")
    
    # Get category
    category = await db.main_categories.find_one({"id": update_data.main_category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Main category not found")
    
    # Calculate the difference in pieces to adjust inventory
    old_pieces_sold = existing_tracking.get("pieces_sold", 0)
    new_pieces_sold = update_data.pieces_sold
    pieces_difference = new_pieces_sold - old_pieces_sold
    
    # If pieces increased, deduct more. If decreased, add back
    if pieces_difference != 0:
        purchases = await db.inventory_purchases.find(
            {"main_category_id": update_data.main_category_id},
            {"_id": 0}
        ).sort("purchase_date", 1 if pieces_difference > 0 else -1).to_list(length=None)
        
        pieces_to_adjust = abs(pieces_difference)
        for purchase in purchases:
            if pieces_to_adjust <= 0:
                break
            
            remaining_pieces = purchase.get("remaining_pieces", 0) or 0
            
            if pieces_difference > 0:  # Need to deduct more pieces
                if remaining_pieces > 0:
                    deduction = min(remaining_pieces, pieces_to_adjust)
                    new_remaining = remaining_pieces - deduction
                    pieces_to_adjust -= deduction
                    
                    await db.inventory_purchases.update_one(
                        {"id": purchase["id"]},
                        {"$set": {"remaining_pieces": new_remaining}}
                    )
            else:  # Need to add back pieces
                total_pieces = purchase.get("total_pieces", 0) or 0
                if remaining_pieces < total_pieces:
                    addition = min(total_pieces - remaining_pieces, pieces_to_adjust)
                    new_remaining = remaining_pieces + addition
                    pieces_to_adjust -= addition
                    
                    await db.inventory_purchases.update_one(
                        {"id": purchase["id"]},
                        {"$set": {"remaining_pieces": new_remaining}}
                    )
    
    # Update tracking record
    tracking_date = update_data.tracking_date if update_data.tracking_date else existing_tracking.get("tracking_date")
    
    await db.daily_pieces_tracking.update_one(
        {"id": tracking_id},
        {"$set": {
            "main_category_id": update_data.main_category_id,
            "main_category_name": category["name"],
            "tracking_date": tracking_date,
            "pieces_sold": new_pieces_sold
        }}
    )
    
    updated_tracking = await db.daily_pieces_tracking.find_one({"id": tracking_id}, {"_id": 0})
    logger.info(f"Daily pieces tracking updated: {tracking_id}")
    return DailyPiecesTracking(**updated_tracking)

@api_router.delete("/daily-pieces-tracking/{tracking_id}")
async def delete_daily_pieces_tracking(tracking_id: str, current_user: User = Depends(get_current_user)):
    # Check if user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can delete daily pieces tracking")
    
    # Get existing tracking
    existing_tracking = await db.daily_pieces_tracking.find_one({"id": tracking_id}, {"_id": 0})
    if not existing_tracking:
        raise HTTPException(status_code=404, detail="Tracking record not found")
    
    # Add back the pieces to inventory
    pieces_to_add_back = existing_tracking.get("pieces_sold", 0)
    
    if pieces_to_add_back > 0:
        purchases = await db.inventory_purchases.find(
            {"main_category_id": existing_tracking["main_category_id"]},
            {"_id": 0}
        ).sort("purchase_date", -1).to_list(length=None)
        
        for purchase in purchases:
            if pieces_to_add_back <= 0:
                break
            
            remaining_pieces = purchase.get("remaining_pieces", 0) or 0
            total_pieces = purchase.get("total_pieces", 0) or 0
            
            if remaining_pieces < total_pieces:
                addition = min(total_pieces - remaining_pieces, pieces_to_add_back)
                new_remaining = remaining_pieces + addition
                pieces_to_add_back -= addition
                
                await db.inventory_purchases.update_one(
                    {"id": purchase["id"]},
                    {"$set": {"remaining_pieces": new_remaining}}
                )
    
    # Delete tracking record
    await db.daily_pieces_tracking.delete_one({"id": tracking_id})
    logger.info(f"Daily pieces tracking deleted: {tracking_id}")
    return {"message": "Pieces tracking deleted successfully"}

# Daily Waste Tracking
@api_router.get("/daily-waste-tracking", response_model=List[DailyWasteTracking])
async def get_daily_waste_tracking(
    main_category_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if main_category_id:
        query["main_category_id"] = main_category_id
    if start_date and end_date:
        query["tracking_date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["tracking_date"] = {"$gte": start_date}
    
    tracking = await db.daily_waste_tracking.find(query, {"_id": 0}).sort("tracking_date", -1).to_list(length=None)
    
    # Handle backwards compatibility - convert old field names to new format
    for record in tracking:
        if 'waste_kg' not in record:
            # Old format had waste_weight_kg
            record['waste_kg'] = record.get('waste_weight_kg', 0)
        # Remove old fields if they exist
        record.pop('waste_weight_kg', None)
        record.pop('raw_weight_kg', None)
        record.pop('dressed_weight_kg', None)
        record.pop('waste_percentage', None)
    
    return tracking

@api_router.post("/daily-waste-tracking", response_model=DailyWasteTracking)
async def create_daily_waste_tracking(tracking: DailyWasteTrackingCreate, current_user: User = Depends(get_current_user)):
    # Get main category
    category = await db.main_categories.find_one({"id": tracking.main_category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Main category not found")
    
    # Validate waste amount
    if tracking.waste_kg <= 0:
        raise HTTPException(status_code=400, detail="Waste weight must be greater than 0")
    
    # Determine tracking date (use IST)
    tracking_date = tracking.tracking_date if tracking.tracking_date else get_ist_now().strftime("%Y-%m-%d")
    
    # Deduct waste weight from inventory using FIFO
    purchases = await db.inventory_purchases.find(
        {"main_category_id": tracking.main_category_id},
        {"_id": 0}
    ).sort("purchase_date", 1).to_list(length=None)
    
    weight_to_deduct = tracking.waste_kg
    for purchase in purchases:
        if weight_to_deduct <= 0:
            break
        
        remaining = purchase.get("remaining_weight_kg", 0)
        if remaining > 0:
            deduction = min(remaining, weight_to_deduct)
            new_remaining = remaining - deduction
            weight_to_deduct -= deduction
            
            await db.inventory_purchases.update_one(
                {"id": purchase["id"]},
                {"$set": {"remaining_weight_kg": round(new_remaining, 2)}}
            )
    
    if weight_to_deduct > 0:
        logger.warning(f"Not enough inventory for {category['name']}. {weight_to_deduct}kg could not be deducted.")
    
    # Create waste tracking record
    new_tracking = DailyWasteTracking(
        main_category_id=tracking.main_category_id,
        main_category_name=category["name"],
        tracking_date=tracking_date,
        waste_kg=round(tracking.waste_kg, 2),
        notes=tracking.notes
    )
    
    await db.daily_waste_tracking.insert_one(new_tracking.dict())
    logger.info(f"Daily waste tracking created: {category['name']} - Waste: {tracking.waste_kg}kg on {tracking_date}")
    return new_tracking

@api_router.put("/daily-waste-tracking/{tracking_id}")
async def update_daily_waste_tracking(
    tracking_id: str,
    update_data: DailyWasteTrackingCreate,
    current_user: User = Depends(get_current_user)
):
    # Check if user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can edit daily waste tracking")
    
    # Get existing tracking
    existing_tracking = await db.daily_waste_tracking.find_one({"id": tracking_id}, {"_id": 0})
    if not existing_tracking:
        raise HTTPException(status_code=404, detail="Tracking record not found")
    
    # Get category
    category = await db.main_categories.find_one({"id": update_data.main_category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Main category not found")
    
    # Validate waste amount
    if update_data.waste_kg <= 0:
        raise HTTPException(status_code=400, detail="Waste weight must be greater than 0")
    
    # Calculate the difference in waste to adjust inventory
    old_waste_kg = existing_tracking.get("waste_kg", 0)
    new_waste_kg = update_data.waste_kg
    waste_difference = new_waste_kg - old_waste_kg
    
    # If waste increased, deduct more. If decreased, add back
    if waste_difference != 0:
        purchases = await db.inventory_purchases.find(
            {"main_category_id": update_data.main_category_id},
            {"_id": 0}
        ).sort("purchase_date", 1 if waste_difference > 0 else -1).to_list(length=None)
        
        weight_to_adjust = abs(waste_difference)
        for purchase in purchases:
            if weight_to_adjust <= 0:
                break
            
            remaining_weight = purchase.get("remaining_weight_kg", 0)
            
            if waste_difference > 0:  # Need to deduct more weight
                if remaining_weight > 0:
                    deduction = min(remaining_weight, weight_to_adjust)
                    new_remaining = remaining_weight - deduction
                    weight_to_adjust -= deduction
                    
                    await db.inventory_purchases.update_one(
                        {"id": purchase["id"]},
                        {"$set": {"remaining_weight_kg": round(new_remaining, 2)}}
                    )
            else:  # Need to add back weight
                total_weight = purchase.get("total_weight_kg", 0)
                if remaining_weight < total_weight:
                    addition = min(total_weight - remaining_weight, weight_to_adjust)
                    new_remaining = remaining_weight + addition
                    weight_to_adjust -= addition
                    
                    await db.inventory_purchases.update_one(
                        {"id": purchase["id"]},
                        {"$set": {"remaining_weight_kg": round(new_remaining, 2)}}
                    )
    
    # Update tracking record
    tracking_date = update_data.tracking_date if update_data.tracking_date else existing_tracking.get("tracking_date")
    
    await db.daily_waste_tracking.update_one(
        {"id": tracking_id},
        {"$set": {
            "main_category_id": update_data.main_category_id,
            "main_category_name": category["name"],
            "tracking_date": tracking_date,
            "waste_kg": round(new_waste_kg, 2),
            "notes": update_data.notes
        }}
    )
    
    updated_tracking = await db.daily_waste_tracking.find_one({"id": tracking_id}, {"_id": 0})
    logger.info(f"Daily waste tracking updated: {tracking_id}")
    return DailyWasteTracking(**updated_tracking)

@api_router.delete("/daily-waste-tracking/{tracking_id}")
async def delete_daily_waste_tracking(tracking_id: str, current_user: User = Depends(get_current_user)):
    # Check if user is admin
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc.get('is_admin', False):
        raise HTTPException(status_code=403, detail="Only admin can delete daily waste tracking")
    
    # Get existing tracking
    existing_tracking = await db.daily_waste_tracking.find_one({"id": tracking_id}, {"_id": 0})
    if not existing_tracking:
        raise HTTPException(status_code=404, detail="Tracking record not found")
    
    # Add back the waste weight to inventory
    weight_to_add_back = existing_tracking.get("waste_kg", 0)
    
    if weight_to_add_back > 0:
        purchases = await db.inventory_purchases.find(
            {"main_category_id": existing_tracking["main_category_id"]},
            {"_id": 0}
        ).sort("purchase_date", -1).to_list(length=None)
        
        for purchase in purchases:
            if weight_to_add_back <= 0:
                break
            
            remaining_weight = purchase.get("remaining_weight_kg", 0)
            total_weight = purchase.get("total_weight_kg", 0)
            
            if remaining_weight < total_weight:
                addition = min(total_weight - remaining_weight, weight_to_add_back)
                new_remaining = remaining_weight + addition
                weight_to_add_back -= addition
                
                await db.inventory_purchases.update_one(
                    {"id": purchase["id"]},
                    {"$set": {"remaining_weight_kg": round(new_remaining, 2)}}
                )
    
    # Delete tracking record
    await db.daily_waste_tracking.delete_one({"id": tracking_id})
    logger.info(f"Daily waste tracking deleted: {tracking_id}")
    return {"message": "Waste tracking deleted successfully"}

# New POS Sales
@api_router.post("/pos-sales", response_model=POSSaleNew)
async def create_pos_sale(sale: POSSaleCreateNew, current_user: User = Depends(get_current_user)):
    # Validate all derived products and main categories
    for item in sale.items:
        # Check derived product exists
        product = await db.derived_products.find_one({"id": item.derived_product_id}, {"_id": 0})
        if not product:
            raise HTTPException(status_code=404, detail=f"Derived product {item.derived_product_id} not found")
        
        # Check main category exists
        category = await db.main_categories.find_one({"id": item.main_category_id}, {"_id": 0})
        if not category:
            raise HTTPException(status_code=404, detail=f"Main category {item.main_category_id} not found")
    
    # Deduct inventory weight/pieces from purchases (FIFO)
    for item in sale.items:
        # Deduct weight if applicable (weight or package units)
        if item.quantity_kg > 0:
            weight_to_deduct = item.quantity_kg
            
            # Get all purchases for this main category, sorted by purchase date (FIFO)
            purchases = await db.inventory_purchases.find(
                {"main_category_id": item.main_category_id},
                {"_id": 0}
            ).sort("purchase_date", 1).to_list(length=None)
            
            for purchase in purchases:
                if weight_to_deduct <= 0:
                    break
                
                remaining = purchase.get("remaining_weight_kg", 0)
                if remaining > 0:
                    deduction = min(remaining, weight_to_deduct)
                    new_remaining = remaining - deduction
                    weight_to_deduct -= deduction
                    
                    await db.inventory_purchases.update_one(
                        {"id": purchase["id"]},
                        {"$set": {"remaining_weight_kg": round(new_remaining, 2)}}
                    )
        
        # Deduct pieces if applicable (pieces unit)
        if item.quantity_pieces and item.quantity_pieces > 0:
            pieces_to_deduct = item.quantity_pieces
            
            # Get all purchases for this main category, sorted by purchase date (FIFO)
            purchases = await db.inventory_purchases.find(
                {"main_category_id": item.main_category_id},
                {"_id": 0}
            ).sort("purchase_date", 1).to_list(length=None)
            
            for purchase in purchases:
                if pieces_to_deduct <= 0:
                    break
                
                remaining_pieces = purchase.get("remaining_pieces", 0) or 0
                if remaining_pieces > 0:
                    deduction = min(remaining_pieces, pieces_to_deduct)
                    new_remaining = remaining_pieces - deduction
                    pieces_to_deduct -= deduction
                    
                    await db.inventory_purchases.update_one(
                        {"id": purchase["id"]},
                        {"$set": {"remaining_pieces": new_remaining}}
                    )
    
    # Create sale record
    new_sale = POSSaleNew(**sale.dict())
    await db.pos_sales.insert_one(new_sale.dict())
    
    # Update customer total purchases if customer provided
    if sale.customer_id:
        await db.customers.update_one(
            {"id": sale.customer_id},
            {"$inc": {"total_purchases": sale.total}}
        )
    
    logger.info(f"POS sale created: Total {sale.total}")
    return new_sale

@api_router.get("/pos-sales", response_model=List[POSSaleNew])
async def get_pos_sales(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    query = {}
    if start_date and end_date:
        query["sale_date"] = {
            "$gte": datetime.fromisoformat(start_date),
            "$lte": datetime.fromisoformat(end_date)
        }
    
    sales = await db.pos_sales.find(query, {"_id": 0}).sort("sale_date", -1).to_list(length=None)
    return sales



@api_router.get("/")
async def root():
    return {"message": "Meat Inventory API"}

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
